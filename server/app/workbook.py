"""Multi-sheet `.xlsx` workbooks (#86) — the artifact an accountant reconciles in.

Export used to be one flat CSV of one view. This module answers the ask recorded
verbatim in #88 ("be able to export it to an Excel sheet") with a workbook whose
sheets are the epic's figures: cost, revenue and fee per CLIN, the #77 rate buildup,
the #80 fee position, and the timesheet and funding rows underneath them so every
dollar above is traceable to something.

Three rules carry over from the Profitability view (#82), and breaking any of them
here would be the same defect in a new medium:

1. **Withholding travels as a reason, never as a zero.** `web/src/profitability.js` is
   the reference implementation and the message strings below are copied from it
   deliberately, so the workbook and the view give a reader the *same sentence* for the
   same refusal. A withheld figure is an em dash carrying a cell comment. A `.xlsx`
   full of zeroes where the engine withheld is a claim the engine refused to make.
2. **Never synthesize a number the engine didn't publish.** A T&M line's at-completion
   cell is empty in the view; it is empty here. Per-person cost/revenue/fee contribution
   is #150 and is not built, so the people sheet carries hours and rates and says so
   rather than multiplying them into a figure nothing else in the app would agree with.
3. **Sums are live formulas.** An accountant changes an assumption and expects the total
   to move; a workbook of frozen values is a screenshot with extra steps. Every total is
   `=SUM(...)`, and the Summary sheet's money cells are references into `By CLIN` — so
   the top-line figure and its backing rows cannot drift apart in the reader's hands.

The clause trap (#81) is sidestepped rather than re-implemented: this module prints the
fee clause *number* the payload carries and never maps it to a name. The two clause maps
that must not merge live in `web/src/format.js`, and a second copy on the server is a
second chance to merge them.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- formats ---------------------------------------------------------------------
#
# "An accountant should not have to reformat a single column" is an acceptance
# criterion, so formats are applied per cell rather than left to Excel's guess.

MONEY = '"$"#,##0.00'
MONEY0 = '"$"#,##0'
PCT = "0.0%"
PCT2 = "0.00%"
RATE = '"$"#,##0.00'
HOURS = "#,##0.0"
DATE_FMT = "yyyy-mm-dd"
STAMP_FMT = "yyyy-mm-dd hh:mm"

DASH = "—"

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
LABEL_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, size=9, color="6B7280")
WITHHELD_FONT = Font(italic=True, color="9CA3AF")
TOTAL_TOP = Border(top=Side(style="thin", color="9CA3AF"))

# ---- the withholding contract ----------------------------------------------------
#
# Ported from web/src/profitability.js. The strings are identical on purpose: a user who
# reads "Cost equals billings at cost-model level 1" in the view and opens the workbook
# should meet the same words, not a paraphrase they have to reconcile.

LEVEL_1_COST = (
    "Cost equals billings at cost-model level 1 — no direct rates have been supplied."
)
LEVEL_1_FEE = (
    "Fee is structural at level 1: it reconciles with cost and revenue but says "
    "nothing about profit."
)
NO_FEE_TERMS = "This award printed no fee figures for the engine to earn against."
NO_FEE_TERMS_CLIN = (
    "This CLIN's award printed no fee figures for the engine to earn against."
)
COST_IS_STANDIN = "Cost is a billing-rate stand-in on this CLIN."
FEE_IS_STANDIN_CLIN = (
    "This CLIN's hours are priced at the billing rate, so its fee reconciles with cost "
    "and revenue but says nothing about profit."
)
PARTIAL_COST = (
    "Part of this contract's hours are priced at the billing rate, so total cost is not "
    "known — every labor category needs a direct rate before a contract margin means "
    "anything."
)
PARTIAL_FEE = (
    "Total fee is revenue less cost, and part of this contract's cost is a billing-rate "
    "stand-in."
)
NO_REVENUE_CLIN = "No revenue recognised on this CLIN yet."
UNSUPPORTED_POLICY = "Contract policy is currently unsupported."
PRICE_NOT_REVENUE = (
    "Fixed-price work earns its price on delivery, and Runway has no milestone or "
    "delivery input — so a contract total counting that price as revenue would report "
    "unstarted work as earned. Each fixed-price CLIN's price and the cost against it "
    "are in its at-completion position below."
)
PRICE_NOT_REVENUE_CLIN = (
    "This CLIN earns its price on delivery (FAR 16.202) and Runway is told about no "
    "deliveries, so none of it is recognised yet. The price is in the ceiling column "
    "and the cost against it in the at-completion position."
)
PRICE_NOT_REVENUE_FEE = (
    "Fee is revenue less cost, and this contract's fixed-price work has no recognised "
    "revenue to take it from — a price not yet earned less the cost so far is unspent "
    "budget, not profit."
)
PRICE_NOT_REVENUE_FEE_CLIN = (
    "Fee is revenue less cost, and this CLIN has no recognised revenue to take it "
    "from — its price less the cost so far is unspent budget, not profit, until it "
    "delivers."
)
PASS_THROUGH = (
    "A non-labor CLIN is a cost-reimbursable pass-through: its logged travel, ODC and "
    "materials dollars consume funding and earn no fee."
)
FEE_NEEDS_COST = (
    "Earned fee is a function of cost, and where cost is the billing rate standing in "
    "there is no fee to read — supply direct rates for this CLIN's categories."
)
FEE_NEEDS_TERMS = (
    "This award's fee terms are incomplete, so everything earned against them is "
    "withheld rather than computed to zero."
)
NO_REVENUE = "No revenue recognised yet."


def fact(value):
    """A figure the workbook may print."""
    return {"value": value, "withheld": None}


def withheld(why: str):
    """A refusal, carrying the reason a reader needs. `value` is None exactly when
    `withheld` is set, so no caller can accidentally format a withheld number."""
    return {"value": None, "withheld": why}


def margin_available(burn: dict) -> bool:
    """`rates.CostModel.margin_available` (rates.py:320) — which tier the contract's
    rate ladder is configured at, and *not* the gate (#152).

    It goes true on the first indirect pool plus the first direct rate, so a contract
    whose ladder covers one of six labor categories reads true while five are still
    billing-rate stand-ins. What a figure may claim is gated on the engine's own cost
    truth instead — `totals.cost_known` and each CLIN's `cost_known` — and this flag
    only chooses which sentence explains a refusal.
    """
    return bool(
        ((burn or {}).get("contract") or {})
        .get("cost_model", {})
        .get("margin_available")
    )


def summary_figures(burn: dict) -> dict:
    """The four contract-level figures, under profitability.js's `summary` rules."""
    t = (burn or {}).get("totals") or {}
    margin = margin_available(burn)
    # A total cost that is part buildup and part billing stand-in is not a contract
    # cost, and a margin taken off it is arithmetic wearing a fact's clothes (#152).
    cost_known = t.get("cost_known") is True
    # Whether that revenue is revenue at all (#154): on fixed-price work the engine
    # reports the contract price in the slot, because the price is earned on delivery
    # and no delivery input exists. Revenue is asked before cost on the fee and margin
    # rows — more rates will not unlock a fee with no recognised revenue under it.
    revenue_known = t.get("revenue_known") is True
    cost_why = PARTIAL_COST if margin else LEVEL_1_COST
    revenue = t.get("revenue") or 0
    cost = t.get("cost") or 0
    return {
        "revenue": fact(revenue) if revenue_known else withheld(PRICE_NOT_REVENUE),
        "cost": fact(cost) if cost_known else withheld(cost_why),
        "fee": (
            withheld(PRICE_NOT_REVENUE_FEE)
            if not revenue_known
            else (
                withheld(PARTIAL_FEE if margin else LEVEL_1_FEE)
                if not cost_known
                else (
                    fact(t.get("fee") or 0)
                    if t.get("fee_known")
                    else withheld(NO_FEE_TERMS)
                )
            )
        ),
        "margin": (
            withheld(PRICE_NOT_REVENUE)
            if not revenue_known
            else (
                fact((revenue - cost) / revenue)
                if cost_known and revenue
                else withheld(cost_why if not cost_known else NO_REVENUE)
            )
        ),
    }


def clin_figures(clin: dict, margin: bool) -> dict:
    """One CLIN's money columns — profitability.js's `clinFigures`.

    Non-labor CLINs carry no `cost` / `revenue` / `fee_earned` keys at all; defaulting
    them to 0 prints "$0" on a line holding real spend. Verified against a live payload,
    not read off the comments.

    A labor CLIN gates on its own `cost_known` and never on the contract's tier (#152):
    a mixed award prices one line from category rates and leaves another on the billing
    fallback, and a contract-wide unlock would print the fallback line's billings under
    a cost heading.
    """
    policy = clin.get("pricing_policy") or {}
    unsupported = policy.get("status") == "unsupported"
    unsupported_notice = policy.get("notice") or UNSUPPORTED_POLICY
    if not clin.get("is_labor"):
        spent = clin.get("spent") or 0
        # Unless the award priced this line fixed (#155): then it is a deliverable, not
        # a pass-through, and the refusals are the fixed-price pair — a price earned on
        # delivery, and a fee that cannot exist before the revenue does. Mirrors
        # `clinFigures` in `web/src/profitability.js`, cell for cell.
        if clin.get("margin_managed"):
            return {
                "revenue": withheld(PRICE_NOT_REVENUE_CLIN),
                "cost": fact(spent),
                "fee": withheld(
                    unsupported_notice if unsupported else PRICE_NOT_REVENUE_FEE_CLIN
                ),
                "margin": withheld(
                    unsupported_notice if unsupported else PRICE_NOT_REVENUE_FEE_CLIN
                ),
            }
        return {
            "revenue": fact(spent),
            "cost": fact(spent),
            "fee": withheld(unsupported_notice if unsupported else PASS_THROUGH),
            "margin": withheld(unsupported_notice if unsupported else PASS_THROUGH),
        }
    cost_known = clin.get("cost_known") is True
    # Per CLIN, never off the contract flag: on a mixed award the T&M line beside a
    # fixed-price one recognises revenue every week and keeps printing it (#154).
    revenue_known = clin.get("revenue_known") is True
    cost_why = COST_IS_STANDIN if margin else LEVEL_1_COST
    margin_pct = clin.get("margin_pct")
    if unsupported:
        return {
            "revenue": (
                fact(clin.get("revenue") or 0)
                if revenue_known
                else withheld(PRICE_NOT_REVENUE_CLIN)
            ),
            "cost": fact(clin.get("cost") or 0) if cost_known else withheld(cost_why),
            "fee": withheld(unsupported_notice),
            "margin": withheld(unsupported_notice),
        }
    return {
        "revenue": (
            fact(clin.get("revenue") or 0)
            if revenue_known
            else withheld(PRICE_NOT_REVENUE_CLIN)
        ),
        "cost": fact(clin.get("cost") or 0) if cost_known else withheld(cost_why),
        "fee": (
            withheld(PRICE_NOT_REVENUE_FEE_CLIN)
            if not revenue_known
            else (
                withheld(FEE_IS_STANDIN_CLIN if margin else LEVEL_1_FEE)
                if not cost_known
                else (
                    fact(clin.get("fee_earned") or 0)
                    if clin.get("fee_known")
                    else withheld(NO_FEE_TERMS_CLIN)
                )
            )
        ),
        # With cost known, a null `margin_pct` is one of the two remaining refusals,
        # and they take different fixes: unstated fee terms are fixed by importing a
        # document, no revenue yet by waiting.
        "margin": (
            withheld(PRICE_NOT_REVENUE_CLIN)
            if not revenue_known
            else (
                withheld(cost_why)
                if not cost_known
                else (
                    fact(margin_pct)
                    if margin_pct is not None
                    else withheld(
                        NO_REVENUE_CLIN if clin.get("fee_known") else NO_FEE_TERMS_CLIN
                    )
                )
            )
        ),
    }


def fee_figures(fp: dict) -> dict:
    """The fee position's figures — profitability.js's `feeFigures`.

    Award-stated facts survive level 1, computed ones don't: `target` and the clause are
    printed on the award and true before an hour is charged, while everything below is a
    function of `cost_frac` and so of the billing rate at level 1.

    Both truth flags default closed (#153). A payload that omits one has not said the
    figure is trustworthy, and unknown *terms* matter as much as unknown cost: fee
    earned against a structure the award never printed computes a clean $0, which is a
    claim that the work has earned nothing.
    """
    why = (
        FEE_NEEDS_TERMS
        if fp.get("terms_known") is not True
        else (None if fp.get("cost_known") is True else FEE_NEEDS_COST)
    )
    gated = (lambda v: withheld(why)) if why else (lambda v: fact(v or 0))
    return {
        "target": (
            withheld("The award printed no fee target.")
            if fp.get("target") is None
            else fact(fp["target"])
        ),
        "earned": gated(fp.get("earned")),
        "at_completion": gated(fp.get("at_completion")),
        "delta": (
            withheld("No fee target to measure against.")
            if fp.get("target_delta") is None
            else gated(fp.get("target_delta"))
        ),
        "at_risk": gated(fp.get("at_risk")),
        "absorbed": gated(fp.get("absorbed")),
        "withhold": gated(fp.get("withhold")),
        "collectable": gated(fp.get("collectable")),
    }


def ordered_clins(burn: dict) -> list:
    """Labor first, then non-labor — the order every other surface lists them in."""
    clins = (burn or {}).get("clins") or []
    return [c for c in clins if c.get("is_labor")] + [
        c for c in clins if not c.get("is_labor")
    ]


POOL_BASE = {
    "direct_labor": "direct labor",
    "labor_plus_fringe": "direct labor + fringe",
    "total_cost_input": "total cost input (labor + fringe + overhead)",
}

COST_SOURCE = {
    "employee_direct": "Per-person direct rate",
    "lcat_direct": "Category (LCAT) direct rate",
    "negotiated_fallback": "Billing rate, standing in for cost",
    "none": "Not priced",
}

MEASURED = {"cost": "cost", "billings": "billings", "price": "price"}


# ---- cell plumbing ---------------------------------------------------------------


def _cell(ws, row, col, value=None, fmt=None, font=None, align=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fmt:
        c.number_format = fmt
    if font:
        c.font = font
    if align:
        c.alignment = Alignment(horizontal=align)
    return c


def _figure(ws, row, col, figure: dict, fmt=MONEY, formula: Optional[str] = None):
    """Write a `{value, withheld}` figure.

    Withheld renders as an em dash carrying its explanation as a cell comment — the one
    thing it must never render as is `0`, which is a number a reader would act on.
    """
    c = ws.cell(row=row, column=col)
    if figure.get("withheld"):
        c.value = DASH
        c.font = WITHHELD_FONT
        c.alignment = Alignment(horizontal="right")
        c.comment = Comment(figure["withheld"], "Runway", height=110, width=320)
    else:
        c.value = formula if formula else figure.get("value")
        c.number_format = fmt
    return c


def _headers(ws, row, labels, widths=None, freeze=True):
    """A header band. `freeze=False` for a sheet's *second* table — freezing there
    would pin every row above it to the top of the window, which is worse than not
    freezing at all."""
    for i, label in enumerate(labels, start=1):
        c = _cell(ws, row, i, label)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _section(ws, row, title):
    c = _cell(ws, row, 1, title, font=SECTION_FONT)
    return c


def _note(ws, row, col, text):
    if text:
        _cell(ws, row, col, text, font=NOTE_FONT)


def _as_date(value):
    """A date cell where the payload carries an ISO string, the string otherwise —
    a half-parsed date printed as a number is worse than the text it came from."""
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return str(value)


def _contract_type(burn: dict) -> str:
    """The contract's type as its labor CLINs price it. 'Mixed' where they disagree —
    a real state on a mixed award, and one label would hide it."""
    labels = []
    for c in ordered_clins(burn):
        p = c.get("pricing_policy") or {}
        label = (
            "Unsupported"
            if p.get("status") == "unsupported"
            else (p.get("label") if p.get("known") else None)
        )
        if label and label not in labels:
            labels.append(label)
    if not labels:
        return "Not stated on the award"
    return labels[0] if len(labels) == 1 else "Mixed: " + ", ".join(labels)


def _provenance(contract_row: dict) -> Optional[str]:
    """The SIMULATED caveat, when the source was generated test data.

    `sync_seed` is set only by a Fixtura sync, so its presence is the honest marker —
    and a workbook that gets emailed to a CO without it is the failure mode.
    """
    seed = (contract_row or {}).get("sync_seed")
    if seed is None:
        return None
    return (
        f"SIMULATED DATA — timesheets for this contract were generated by Fixtura "
        f"(seed {seed}), not extracted from an accounting system. Every figure derived "
        f"from hours below is test data."
    )


# ---- sheets ----------------------------------------------------------------------


def _clin_sheet(ws, burn: dict, margin: bool) -> dict:
    """`By CLIN` — one row per CLIN, and the totals every other sheet points at."""
    cols = [
        "CLIN",
        "Name",
        "Pricing policy",
        "Spend measured in",
        "Ceiling",
        "Funded",
        "Spent",
        "Cost",
        "Revenue",
        "Fee earned",
        "Margin %",
        "Weekly burn",
        "Runway (days)",
        "Status",
    ]
    _headers(ws, 1, cols, [12, 34, 24, 16, 14, 14, 14, 14, 14, 14, 10, 14, 13, 22])
    row = 2
    first = row
    for c in ordered_clins(burn):
        figs = clin_figures(c, margin)
        policy = c.get("pricing_policy") or {}
        _cell(ws, row, 1, c.get("code") or c.get("id"))
        _cell(ws, row, 2, c.get("name"))
        policy_cell = _cell(
            ws,
            row,
            3,
            (
                "Unsupported"
                if policy.get("status") == "unsupported"
                else policy.get("label") if policy.get("known") else "Not stated"
            ),
        )
        if policy.get("status") == "unsupported":
            policy_cell.comment = Comment(
                policy.get("notice") or UNSUPPORTED_POLICY, "Runway"
            )
        _cell(
            ws,
            row,
            4,
            MEASURED.get(c.get("measured_against"), c.get("measured_against")),
        )
        _cell(ws, row, 5, c.get("ceiling"), MONEY)
        _cell(ws, row, 6, c.get("funded"), MONEY)
        _cell(ws, row, 7, c.get("spent"), MONEY)
        _figure(ws, row, 8, figs["cost"])
        _figure(ws, row, 9, figs["revenue"])
        _figure(ws, row, 10, figs["fee"])
        # Live rather than frozen: fee over revenue, the same definition as the engine's
        # `margin_pct`, so an accountant who edits a cost cell watches the margin move
        # instead of watching a stale percentage sit still.
        _figure(
            ws,
            row,
            11,
            figs["margin"],
            PCT,
            formula=f'=IF(I{row}=0,"",(I{row}-H{row})/I{row})',
        )
        _cell(ws, row, 12, c.get("weekly"), MONEY)
        _cell(ws, row, 13, c.get("runway_days"))
        _cell(ws, row, 14, c.get("status_label") or c.get("status"))
        row += 1
    last = row - 1
    totals = summary_figures(burn)
    _cell(ws, row, 1, "Total", font=LABEL_FONT)
    for col in (5, 6, 7, 12):
        letter = get_column_letter(col)
        c = _cell(ws, row, col, f"=SUM({letter}{first}:{letter}{last})", MONEY)
        c.font = LABEL_FONT
    # The contract-level figure governs its own total cell: where the engine withheld
    # cost, a `=SUM()` over a column of em dashes would quietly report $0.
    for col, key in ((8, "cost"), (9, "revenue"), (10, "fee")):
        letter = get_column_letter(col)
        c = _figure(
            ws,
            row,
            col,
            totals[key],
            MONEY,
            formula=f"=SUM({letter}{first}:{letter}{last})",
        )
        if not totals[key].get("withheld"):
            c.font = LABEL_FONT
    _figure(
        ws,
        row,
        11,
        totals["margin"],
        PCT,
        formula=f'=IF(I{row}=0,"",(I{row}-H{row})/I{row})',
    )
    for col in range(1, len(cols) + 1):
        ws.cell(row=row, column=col).border = TOTAL_TOP
    return {"total_row": row, "first": first, "last": last}


def _summary_sheet(
    ws, burn: dict, contract_row: dict, funding: dict, refs: dict, generated_at
):
    """`Summary` — identity, the money, and where the money came from.

    Every dollar here is a reference into `By CLIN` rather than a copy of the payload,
    which is how the acceptance criterion "every dollar figure on the Summary sheet is
    reproducible from a later sheet" stays true after the reader edits something.
    """
    c = (burn or {}).get("contract") or {}
    t = (burn or {}).get("totals") or {}
    hero = (burn or {}).get("hero") or {}
    sync = (burn or {}).get("sync") or {}
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 74
    tr = refs["total_row"]

    _cell(ws, 1, 1, "RUNWAY EXPORT", font=TITLE_FONT)
    _cell(
        ws,
        1,
        3,
        "Generated by Runway. Totals are live formulas — edit a cell on a "
        "later sheet and the figures here follow.",
        font=NOTE_FONT,
    )
    row = 3
    _section(ws, row, "CONTRACT")
    row += 1
    identity = [
        ("Name", c.get("name")),
        ("PIID", c.get("piid")),
        ("Legal name", c.get("legal_name")),
        ("Agency", c.get("agency")),
        ("Contracting officer", c.get("contracting_officer")),
        ("COR", c.get("cor")),
        ("Vehicle", c.get("vehicle")),
        ("Contract type", _contract_type(burn)),
    ]
    for label, value in identity:
        _cell(ws, row, 1, label, font=LABEL_FONT)
        _cell(ws, row, 2, value if value is not None else DASH)
        row += 1
    for label, value in (
        ("PoP start", c.get("pop_start")),
        ("PoP end", c.get("pop_end")),
    ):
        _cell(ws, row, 1, label, font=LABEL_FONT)
        _cell(ws, row, 2, _as_date(value), DATE_FMT)
        row += 1
    _cell(ws, row, 1, "Weeks remaining", font=LABEL_FONT)
    _cell(ws, row, 2, c.get("weeks_remaining"))
    row += 2

    _section(ws, row, "MONEY")
    row += 1
    margin = margin_available(burn)
    figs = summary_figures(burn)
    money_rows = [
        (
            "Ceiling",
            fact(t.get("ceiling")),
            MONEY,
            f"='By CLIN'!E{tr}",
            "Sum of CLIN ceilings.",
        ),
        (
            "Funded / obligated",
            fact(c.get("obligated")),
            MONEY,
            None,
            "Obligated on the award and every ingested SF-30. See Funding history.",
        ),
        (
            "Spent",
            fact(t.get("spent")),
            MONEY,
            f"='By CLIN'!G{tr}",
            "Sum of CLIN spend. Each CLIN states the quantity it is measured in.",
        ),
        (
            "Cost",
            figs["cost"],
            MONEY,
            f"='By CLIN'!H{tr}",
            "Direct labour burdened through the rate buildup, plus logged non-labor actuals.",
        ),
        (
            "Revenue",
            figs["revenue"],
            MONEY,
            f"='By CLIN'!I{tr}",
            "Recognised from each CLIN's pricing policy, not from the cost ladder.",
        ),
        (
            "Fee earned",
            figs["fee"],
            MONEY,
            f"='By CLIN'!J{tr}",
            "See Fee position for what the award promised and what is at risk.",
        ),
        (
            "Margin %",
            figs["margin"],
            PCT,
            f'=IF(B{row + 4}=0,"",(B{row + 4}-B{row + 3})/B{row + 4})',
            "Revenue less cost, over revenue — the same definition as the engine's per-CLIN margin.",
        ),
    ]
    for label, figure, fmt, formula, note in money_rows:
        _cell(ws, row, 1, label, font=LABEL_FONT)
        _figure(ws, row, 2, figure, fmt, formula=formula)
        _note(ws, row, 3, figure.get("withheld") or note)
        row += 1
    _cell(ws, row, 1, "Runway (days)", font=LABEL_FONT)
    # A dash rather than a blank: an empty cell in a money block reads as an oversight,
    # and a contract past its PoP or with nothing charged genuinely has no clock.
    _cell(ws, row, 2, hero.get("days") if hero.get("days") is not None else DASH)
    _note(
        ws,
        row,
        3,
        "Measured from the newest timesheet week, not from today — see "
        "'Timesheets as of' below.",
    )
    row += 1
    _cell(ws, row, 1, "Stop date", font=LABEL_FONT)
    _cell(ws, row, 2, _as_date(hero.get("stop_date")) or DASH, DATE_FMT)
    _note(
        ws,
        row,
        3,
        f"Limited by {hero.get('limited_by') or 'unknown'}"
        + (f" on {hero.get('clin')}" if hero.get("clin") else "")
        + (
            f"; funding clause {hero.get('funding_clause')}"
            if hero.get("funding_clause")
            else ""
        ),
    )
    row += 2

    _section(ws, row, "PROVENANCE")
    row += 1
    # Why the money block above carries dashes. Two states reach it and they take
    # different fixes (#152): no rate ladder at all, or one that doesn't reach every
    # labor category — and the second is the easier to mistake for a broken export,
    # because the contract *is* at level 2 and most of the workbook is populated.
    if figs["cost"].get("withheld"):
        level = ((burn or {}).get("contract") or {}).get("cost_model") or {}
        _cell(ws, row, 1, "Cost model", font=LABEL_FONT)
        _cell(ws, row, 2, f"Level {level.get('level', 1)}")
        _note(
            ws,
            row,
            3,
            (
                PARTIAL_COST + " The CLINs that are fully priced keep their own cost "
                "and margin on By CLIN."
                if margin
                else LEVEL_1_COST + " Level 1 is a supported, complete state; one "
                "report is withheld, not the app."
            ),
        )
        row += 1
    stamp = [
        ("Generated at", generated_at, STAMP_FMT),
        (
            "Timesheets as of",
            _as_date(sync.get("as_of") or sync.get("latest_week")),
            DATE_FMT,
        ),
        ("Timesheet rows", sync.get("rows"), None),
        ("People", sync.get("people"), None),
        ("Weeks synced", sync.get("weeks"), None),
        ("Data age (days)", sync.get("data_age_days"), None),
    ]
    for label, value, fmt in stamp:
        _cell(ws, row, 1, label, font=LABEL_FONT)
        _cell(ws, row, 2, value if value is not None else DASH, fmt)
        row += 1
    caveat = _provenance(contract_row)
    if caveat:
        c2 = _cell(ws, row, 1, caveat, font=Font(bold=True, color="B45309"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 30


def _people_sheet(ws, alloc: dict):
    """`By person` — the allocation grid with the rates that priced it.

    Deliberately *not* per-person cost, revenue and fee: that is #150, split out of #82
    because it is the sensitive surface, and the engine publishes no such figure. Hours
    times a rate would be a number nothing else in the app would agree with.
    """
    cols = [
        "Person",
        "Employee ID",
        "CLIN",
        "Labor category",
        "Hours / week",
        "Billing rate",
        "Direct rate",
        "Cost rate",
        "Priced by",
        "Rate matched",
    ]
    _headers(ws, 1, cols, [24, 14, 12, 30, 12, 13, 13, 13, 30, 14])
    row = 2
    clin_names = {
        c.get("id"): c.get("code") or c.get("id") for c in (alloc.get("clins") or [])
    }
    for emp in alloc.get("employees") or []:
        for clin_id, cell in (emp.get("cells") or {}).items():
            if not cell:
                continue
            line = cell.get("rate_line") or {}
            _cell(ws, row, 1, emp.get("name"))
            _cell(ws, row, 2, emp.get("id"))
            _cell(ws, row, 3, clin_names.get(clin_id, clin_id))
            _cell(ws, row, 4, cell.get("lcat") or emp.get("lcat"))
            _cell(ws, row, 5, cell.get("hours"), HOURS)
            _cell(ws, row, 6, cell.get("rate"), RATE)
            _cell(ws, row, 7, line.get("direct"), RATE)
            _cell(ws, row, 8, cell.get("cost_rate"), RATE)
            _cell(
                ws,
                row,
                9,
                COST_SOURCE.get(cell.get("cost_source"), cell.get("cost_source")),
            )
            _cell(
                ws, row, 10, "no — unmatched LCAT" if cell.get("unmatched") else "yes"
            )
            row += 1
    if row > 2:
        _cell(ws, row, 4, "Total hours / week", font=LABEL_FONT)
        c = _cell(ws, row, 5, f"=SUM(E2:E{row - 1})", HOURS)
        c.font = LABEL_FONT
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).border = TOTAL_TOP
        row += 1
    row += 1
    _note(
        ws,
        row,
        1,
        "Hours are the allocation grid's hrs/week for the active period. "
        "Per-person cost, revenue and fee contribution is not published by the engine "
        "(#150) and is deliberately not derived here.",
    )


def _buildup_sheet(ws, burn: dict):
    """`Rate buildup` — the #77 chain, shown as a multiplication a reader can follow.

    The point of the sheet is reconciliation, so each pool states the rate *and the base
    it applied to*, and the worked column is live: change the $100 of direct labour in
    E3 and the whole ladder repositions.
    """
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    cm = ((burn or {}).get("contract") or {}).get("cost_model") or {}
    rate_set = cm.get("rate_set") or {}
    pools = rate_set.get("pools") or []
    if not pools:
        _cell(ws, 1, 1, "No indirect rate buildup", font=SECTION_FONT)
        _note(
            ws,
            2,
            1,
            LEVEL_1_COST + " There is no chain to show at level 1; the "
            "Indirect Rates view is where a buildup is entered.",
        )
        return
    _cell(ws, 1, 1, "INDIRECT RATE BUILDUP", font=TITLE_FONT)
    meta = [
        ("Fiscal year", rate_set.get("fiscal_year")),
        ("Scope", rate_set.get("scope")),
        ("Status", rate_set.get("status")),
        (
            "Chain complete",
            (
                "yes"
                if rate_set.get("complete") is not False
                else "no — a pool the buildup needs is missing"
            ),
        ),
    ]
    row = 2
    for label, value in meta:
        _cell(ws, row, 1, label, font=LABEL_FONT)
        _cell(ws, row, 2, value if value is not None else DASH)
        row += 1
    if rate_set.get("status") != "final":
        _note(
            ws,
            row,
            1,
            "Provisional rates: every figure derived from them is repriced "
            "at the year-end true-up.",
        )
        row += 1
    row += 1

    head = row
    _headers(
        ws,
        head,
        ["Pool", "Rate", "Applies to", "Status", "Amount", "Running cost"],
        [26, 12, 40, 14, 16, 18],
        freeze=False,
    )
    row = head + 1
    base_row = row
    _cell(ws, row, 1, "Direct labor", font=LABEL_FONT)
    _cell(ws, row, 3, "the base of the chain — edit the amount to reprice it")
    _cell(ws, row, 5, 100, MONEY)
    _cell(ws, row, 6, f"=E{row}", MONEY)
    # Which running-cost cell each base points at, filled as the chain is written, so a
    # pool applying to "total cost input" multiplies the cost *after* overhead rather
    # than whatever row happened to precede it.
    base_cells = {"direct_labor": row}
    row += 1
    for p in pools:
        _cell(ws, row, 1, p.get("label") or p.get("name"))
        _cell(ws, row, 2, p.get("rate"), PCT2)
        _cell(ws, row, 3, POOL_BASE.get(p.get("base"), p.get("base") or DASH))
        _cell(ws, row, 4, p.get("status"))
        src = base_cells.get(p.get("base"), row - 1)
        _cell(ws, row, 5, f"=B{row}*F{src}", MONEY)
        _cell(ws, row, 6, f"=F{row - 1}+E{row}", MONEY)
        if p.get("base") == "direct_labor":
            base_cells["labor_plus_fringe"] = row
        elif p.get("base") == "labor_plus_fringe":
            base_cells["total_cost_input"] = row
        row += 1
    _cell(ws, row, 1, "Fully burdened cost", font=LABEL_FONT)
    c = _cell(ws, row, 6, f"=F{row - 1}", MONEY)
    c.font = LABEL_FONT
    # The wrap rate goes in the Rate column, where a reader looking for a multiplier
    # will look for it — a ratio under an "Amount" header reads as dollars.
    _cell(ws, row, 2, f"=F{row - 1}/E{base_row}", "0.000", font=LABEL_FONT)
    _cell(ws, row, 3, "wrap rate on the direct labor above", font=NOTE_FONT)
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = TOTAL_TOP
    row += 3

    variance = [
        dict(v, code=c.get("code"))
        for c in ordered_clins(burn)
        for v in (c.get("rate_variance") or [])
    ]
    if not variance:
        return
    _cell(ws, row, 1, "DERIVED VS NEGOTIATED, BY LABOR CATEGORY", font=SECTION_FONT)
    _note(
        ws,
        row + 1,
        1,
        "Only categories whose cost was actually derived appear — "
        "comparing a billing-rate fallback against itself always reports zero.",
    )
    row += 2
    head = row
    _headers(
        ws,
        head,
        [
            "Labor category",
            "CLIN",
            "Derived cost",
            "Fee rate",
            "Derived price",
            "Negotiated rate",
            "Delta",
            "Delta %",
        ],
        [30, 12, 14, 12, 14, 16, 12, 10],
        freeze=False,
    )
    row = head + 1
    for v in variance:
        _cell(ws, row, 1, v.get("lcat"))
        _cell(ws, row, 2, v.get("code"))
        _cell(ws, row, 3, v.get("derived_cost"), RATE)
        _cell(ws, row, 4, v.get("fee_rate"), PCT2)
        _cell(ws, row, 5, v.get("derived_price"), RATE)
        _cell(ws, row, 6, v.get("negotiated_rate"), RATE)
        _cell(ws, row, 7, f"=F{row}-E{row}", RATE)
        _cell(ws, row, 8, f'=IF(E{row}=0,"",G{row}/E{row})', PCT2)
        row += 1


def _fee_sheet(ws, burn: dict):
    """`Fee position` — what the award promised, what the work earned, what is at risk.

    Fee clause numbers are printed, never named: the two clause maps that must stay
    apart (#81) live in the frontend, and a lookup here would be a second place to merge
    them.
    """
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 72
    row = 1
    _cell(ws, row, 1, "FEE POSITION", font=TITLE_FONT)
    row += 2
    fee_clins = [c for c in ordered_clins(burn) if c.get("fee_position")]
    if not fee_clins:
        _cell(
            ws,
            row,
            1,
            "No CLIN on this contract carries a fee mechanic.",
            font=LABEL_FONT,
        )
        _note(
            ws,
            row + 1,
            1,
            "Fixed-price lines carry a margin position instead, T&M "
            "keeps its fee inside the billing rate, and an unlabelled award has "
            "neither. An empty sheet is the ordinary state on most contracts.",
        )
        return
    for clin in fee_clins:
        fp = clin.get("fee_position") or {}
        _cell(ws, row, 1, f"{clin.get('code')} — {clin.get('name')}", font=SECTION_FONT)
        row += 1
        _cell(ws, row, 1, "Basis", font=LABEL_FONT)
        _cell(ws, row, 2, fp.get("basis"))
        row += 1
        if fp.get("clause"):
            _cell(ws, row, 1, "Fee clause", font=LABEL_FONT)
            _cell(ws, row, 2, fp.get("clause"))
            row += 1
        if not fp.get("terms_known"):
            missing = ", ".join(fp.get("missing") or [])
            _note(
                ws,
                row,
                1,
                (
                    (
                        f"This award's fee terms are incomplete — missing {missing}. "
                        "Import the fee structure to price it."
                    )
                    if missing
                    else "This award printed no fee structure for the engine to earn against."
                ),
            )
            row += 1
        figs = fee_figures(fp)
        lines = [
            ("Fee target (award-stated)", figs["target"], MONEY),
            ("Earned to date", figs["earned"], MONEY),
            ("Projected at completion", figs["at_completion"], MONEY),
            ("Delta vs target", figs["delta"], MONEY),
            ("At risk", figs["at_risk"], MONEY),
            ("Absorbed (overrun eating fee)", figs["absorbed"], MONEY),
            ("Withheld under 52.216-8", figs["withhold"], MONEY),
            ("Collectable now", figs["collectable"], MONEY),
        ]
        for label, figure, fmt in lines:
            _cell(ws, row, 1, label, font=LABEL_FONT)
            _figure(ws, row, 2, figure, fmt)
            if figure.get("withheld"):
                _note(ws, row, 3, figure["withheld"])
            row += 1
        if fp.get("cost_frac") is not None:
            _cell(ws, row, 1, "Cost incurred, share of estimate", font=LABEL_FONT)
            _cell(ws, row, 2, fp.get("cost_frac"), PCT)
            _note(ws, row, 3, "The fraction earned fee is computed from.")
            row += 1

        if fp.get("basis") == "base_plus_award":
            periods = fp.get("periods") or []
            recorded = (fp.get("periods_total") or 0) > 0 or bool(periods)
            _cell(ws, row, 1, "Award fee pool", font=LABEL_FONT)
            _cell(ws, row, 2, fp.get("award_pool"), MONEY)
            row += 1
            _cell(ws, row, 1, "Award fee earned", font=LABEL_FONT)
            _cell(ws, row, 2, fp.get("award_earned"), MONEY)
            row += 1
            _cell(ws, row, 1, "Award fee available", font=LABEL_FONT)
            _cell(ws, row, 2, fp.get("award_available"), MONEY)
            # $0 available with no periods recorded means the pool is unallocated, not
            # spent. Verified on a live CPAF (contract 42) — without this sentence the
            # cell reads as "fee gone".
            _note(
                ws,
                row,
                3,
                (
                    "The pool is unallocated: no evaluation periods have been "
                    "recorded, so there is nothing to earn against yet."
                    if not recorded
                    else f"{fp.get('periods_determined') or 0} of {fp.get('periods_total') or 0} "
                    "evaluation periods determined."
                ),
            )
            row += 1
            if periods:
                head = row
                for i, label in enumerate(
                    [
                        "Evaluation period",
                        "Start",
                        "End",
                        "Pool share ($)",
                        "Status",
                        "Determined amount",
                        "Score",
                    ],
                    start=1,
                ):
                    c = _cell(ws, head, i, label)
                    c.fill = HEAD_FILL
                    c.font = HEAD_FONT
                row += 1
                for p in periods:
                    determined = p.get("status") == "determined"
                    _cell(ws, row, 1, p.get("name"))
                    _cell(ws, row, 2, _as_date(p.get("start")), DATE_FMT)
                    _cell(ws, row, 3, _as_date(p.get("end")), DATE_FMT)
                    # Dollars, not a fraction: an award-fee plan that names its
                    # periods without pricing them splits the pool evenly
                    # (`pricing.py:887`), so a share is money. The view fixed the
                    # same mis-format in d632fca.
                    _cell(ws, row, 4, p.get("pool_share"), MONEY)
                    _cell(ws, row, 5, p.get("status"))
                    # A pending period's amount is not money yet, and a zero there would
                    # read as a determination of zero — which is a different fact.
                    _figure(
                        ws,
                        row,
                        6,
                        (
                            fact(p.get("determined_amount"))
                            if determined
                            else withheld(
                                "This period is pending: the government has "
                                "not determined it, so there is no amount."
                            )
                        ),
                    )
                    _cell(ws, row, 7, p.get("score"))
                    row += 1

        share_raw = fp.get("share_raw")
        if fp.get("basis") in ("incentive_fee", "incentive_profit"):
            _cell(ws, row, 1, "Contractor share", font=LABEL_FONT)
            _cell(ws, row, 2, fp.get("share_contractor"), PCT)
            _note(
                ws, row, 3, f"As the award printed it: {share_raw}" if share_raw else ""
            )
            row += 1
            _cell(ws, row, 1, "Point of total assumption", font=LABEL_FONT)
            if fp.get("pta") is None:
                _figure(
                    ws,
                    row,
                    2,
                    withheld("No price ceiling on the award to compute a PTA from."),
                )
            else:
                _cell(ws, row, 2, fp.get("pta"), MONEY)
                _note(
                    ws,
                    row,
                    3,
                    "Above this cost the contractor absorbs every additional dollar.",
                )
            row += 1
        row += 2


def _timesheet_sheet(ws, rows: list):
    """`Timesheet detail` — the charged hours every figure above resolves to."""
    cols = [
        "Employee",
        "Employee ID",
        "Week ending",
        "Charge code",
        "Labor category",
        "Regular hrs",
        "OT hrs",
        "Holiday hrs",
        "Leave hrs",
        "Total hrs",
        "Paid hrs",
        "Synced at",
    ]
    _headers(ws, 1, cols, [24, 14, 13, 13, 30, 12, 10, 12, 11, 11, 11, 18])
    row = 2
    for r in rows or []:
        _cell(ws, row, 1, r.get("employee"))
        _cell(ws, row, 2, r.get("employee_id"))
        _cell(ws, row, 3, _as_date(r.get("week_ending")), DATE_FMT)
        _cell(ws, row, 4, r.get("charge_code"))
        _cell(ws, row, 5, r.get("labor_category"))
        _cell(ws, row, 6, r.get("reg_hours"), HOURS)
        _cell(ws, row, 7, r.get("ot_hours"), HOURS)
        _cell(ws, row, 8, r.get("holiday_hours"), HOURS)
        _cell(ws, row, 9, r.get("leave_hours"), HOURS)
        _cell(ws, row, 10, r.get("total_hours"), HOURS)
        _cell(ws, row, 11, r.get("paid_hours"), HOURS)
        _cell(ws, row, 12, r.get("synced_at"))
        row += 1
    if row > 2:
        _cell(ws, row, 5, "Total", font=LABEL_FONT)
        for col in range(6, 12):
            letter = get_column_letter(col)
            c = _cell(ws, row, col, f"=SUM({letter}2:{letter}{row - 1})", HOURS)
            c.font = LABEL_FONT
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).border = TOTAL_TOP


def _funding_sheet(ws, funding: dict):
    """`Funding history` — obligations by date, with the running total spelled out.

    The workbook computes its own running total beside the cumulative figure each SF-30
    stated, because those two disagreeing is a real and recurring defect (#128/#129) and
    the person best placed to notice is the one reconciling the contract.
    """
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    top = [
        ("Contract ceiling", funding.get("total_ceiling")),
        ("Total obligated", funding.get("total_obligated")),
        (
            "Incrementally funded",
            "yes" if funding.get("incrementally_funded") else "no",
        ),
    ]
    row = 1
    _cell(ws, row, 1, "FUNDING HISTORY", font=TITLE_FONT)
    row += 1
    for label, value in top:
        _cell(ws, row, 1, label, font=LABEL_FONT)
        if isinstance(value, (int, float)):
            _cell(ws, row, 2, value, MONEY)
        else:
            _cell(ws, row, 2, value if value is not None else DASH)
        row += 1
    row += 1
    head = row
    _headers(
        ws,
        head,
        [
            "Modification",
            "Action",
            "Effective date",
            "Amount",
            "Cumulative obligated (as stated)",
            "Running total (computed)",
            "Description",
        ],
        [16, 40, 14, 16, 22, 22, 60],
    )
    row = head + 1
    first = row
    for h in funding.get("obligation_history") or []:
        _cell(ws, row, 1, h.get("mod"))
        _cell(ws, row, 2, h.get("action"))
        _cell(ws, row, 3, _as_date(h.get("date")), DATE_FMT)
        _cell(ws, row, 4, h.get("amount"), MONEY)
        _cell(ws, row, 5, h.get("cumulative_obligated"), MONEY)
        _cell(ws, row, 6, f"=SUM($D${first}:D{row})", MONEY)
        _cell(ws, row, 7, h.get("description"))
        row += 1
    if row > first:
        _cell(ws, row, 2, "Total obligated", font=LABEL_FONT)
        c = _cell(ws, row, 4, f"=SUM(D{first}:D{row - 1})", MONEY)
        c.font = LABEL_FONT
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = TOTAL_TOP


# ---- entry points ----------------------------------------------------------------


def build_contract_workbook(
    burn: dict,
    alloc: dict,
    timesheets: list,
    funding: dict,
    contract_row: dict,
    generated_at: Optional[datetime] = None,
) -> Workbook:
    """The per-contract workbook: seven sheets, Summary first."""
    generated_at = generated_at or datetime.now()
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    by_clin = wb.create_sheet("By CLIN")
    people = wb.create_sheet("By person")
    buildup = wb.create_sheet("Rate buildup")
    fee = wb.create_sheet("Fee position")
    timesheet = wb.create_sheet("Timesheet detail")
    funding_ws = wb.create_sheet("Funding history")

    refs = _clin_sheet(by_clin, burn, margin_available(burn))
    _summary_sheet(summary, burn, contract_row, funding, refs, generated_at)
    _people_sheet(people, alloc or {})
    _buildup_sheet(buildup, burn)
    _fee_sheet(fee, burn)
    _timesheet_sheet(timesheet, timesheets or [])
    _funding_sheet(funding_ws, funding or {})
    return wb


def build_portfolio_workbook(
    entries: list, generated_at: Optional[datetime] = None
) -> Workbook:
    """The portfolio workbook: one row per contract, same withholding rules.

    `entries` is `[(contract_row, burn)]`. A contract at level 1 contributes dashes in
    the cost, fee and margin columns rather than zeroes, so a portfolio total is never
    the sum of some contracts' real cost and other contracts' silence.
    """
    generated_at = generated_at or datetime.now()
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"
    ws.column_dimensions["A"].width = 30
    _cell(ws, 1, 1, "RUNWAY PORTFOLIO EXPORT", font=TITLE_FONT)
    _cell(ws, 1, 4, generated_at, STAMP_FMT)
    cols = [
        "Contract",
        "PIID",
        "Agency",
        "Type",
        "Ceiling",
        "Obligated",
        "Spent",
        "Cost",
        "Revenue",
        "Fee earned",
        "Margin %",
        "Runway (days)",
        "Status",
        "Timesheets as of",
        "Source",
    ]
    _headers(ws, 3, cols, [30, 18, 22, 24, 14, 14, 14, 14, 14, 14, 10, 13, 22, 15, 34])
    row = 4
    first = row
    for contract_row, burn in entries:
        c = (burn or {}).get("contract") or {}
        t = (burn or {}).get("totals") or {}
        hero = (burn or {}).get("hero") or {}
        sync = (burn or {}).get("sync") or {}
        figs = summary_figures(burn)
        _cell(ws, row, 1, c.get("name"))
        _cell(ws, row, 2, c.get("piid"))
        _cell(ws, row, 3, c.get("agency"))
        _cell(ws, row, 4, _contract_type(burn))
        _cell(ws, row, 5, t.get("ceiling"), MONEY)
        _cell(ws, row, 6, c.get("obligated"), MONEY)
        _cell(ws, row, 7, t.get("spent"), MONEY)
        _figure(ws, row, 8, figs["cost"])
        _figure(ws, row, 9, figs["revenue"])
        _figure(ws, row, 10, figs["fee"])
        _figure(
            ws,
            row,
            11,
            figs["margin"],
            PCT,
            formula=f'=IF(I{row}=0,"",(I{row}-H{row})/I{row})',
        )
        _cell(ws, row, 12, hero.get("days"))
        _cell(ws, row, 13, hero.get("status"))
        _cell(
            ws,
            row,
            14,
            _as_date(sync.get("as_of") or sync.get("latest_week")),
            DATE_FMT,
        )
        _cell(
            ws,
            row,
            15,
            "SIMULATED (Fixtura)" if _provenance(contract_row) else "imported",
        )
        row += 1
    if row > first:
        _cell(ws, row, 1, "Total", font=LABEL_FONT)
        for col in (5, 6, 7, 8, 9, 10):
            letter = get_column_letter(col)
            c = _cell(ws, row, col, f"=SUM({letter}{first}:{letter}{row - 1})", MONEY)
            c.font = LABEL_FONT
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).border = TOTAL_TOP
        row += 1
    row += 1
    _note(
        ws,
        row,
        1,
        "A dash is a figure the engine withheld, not a zero — hover it for "
        "the reason. Column totals sum only the contracts that published the figure.",
    )
    return wb


def to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
